"""
Excel Combiner Module
Combines data from two Excel sheets into a template file.
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os


def combine_excel_sheets(sheet1_path, sheet2_path, template_path, output_path):
    """
    Combine data from two Excel sheets into a template.
    
    Parameters:
    -----------
    sheet1_path : str
        Path to the first Excel file (Project information)
    sheet2_path : str
        Path to the second Excel file (Reference data with quantities)
    template_path : str
        Path to the template Excel file
    output_path : str
        Path where the combined Excel file will be saved
    
    Returns:
    --------
    str : Path to the output file if successful, None if failed
    """
    
    try:
        # Load the template
        print(f"Loading template from: {template_path}")
        wb_template = load_workbook(template_path)
        ws_template = wb_template.active
        
        # Load sheet 1 (Project Information)
        print(f"Loading sheet 1 from: {sheet1_path}")
        wb_sheet1 = load_workbook(sheet1_path)
        ws_sheet1 = wb_sheet1.active
        
        # Load sheet 2 (Reference data)
        print(f"Loading sheet 2 from: {sheet2_path}")
        wb_sheet2 = load_workbook(sheet2_path)
        ws_sheet2 = wb_sheet2.active
        
        # Copy data from sheet 1 (columns A to G, starting from row 2)
        print("Copying data from sheet 1 (Project Information)...")
        sheet1_data = []
        for row_idx, row in enumerate(ws_sheet1.iter_rows(min_row=2), start=2):
            # Get cells, not just values, so we can access hyperlinks
            row_cells = list(row[:7])
            if any(cell.value is not None for cell in row_cells):  # Check if row has data
                sheet1_data.append(row_cells)
                # Write to template starting at row 2
                for col_idx, cell in enumerate(row_cells, start=1):
                    # For column G (7), preserve hyperlinks
                    if col_idx == 7 and cell.hyperlink:
                        ws_template.cell(row=row_idx, column=col_idx).value = cell.value
                        ws_template.cell(row=row_idx, column=col_idx).hyperlink = cell.hyperlink
                    else:
                        # For other columns, paste only the value (preserving template formatting)
                        ws_template.cell(row=row_idx, column=col_idx).value = cell.value
        
        print(f"Copied {len(sheet1_data)} rows from sheet 1")
        
        # Get headers from sheet 2 (row 1, columns B onwards)
        print("Reading headers from sheet 2...")
        sheet2_headers = []
        for col_idx, cell in enumerate(ws_sheet2[1], start=1):
            if col_idx > 1 and cell.value:  # Skip column A (File Reference), get B onwards
                sheet2_headers.append((col_idx, cell.value))
        
        # Write sheet2 headers to template starting at column I (column 9)
        template_col_start = 9  # Column I
        for idx, (_, header) in enumerate(sheet2_headers):
            col_letter = get_column_letter(template_col_start + idx)
            # Paste only the value to preserve template formatting
            ws_template[f"{col_letter}1"].value = header
        
        print(f"Added {len(sheet2_headers)} headers from sheet 2")
        
        # Create a mapping of File Reference to row data from sheet 2
        print("Creating reference mapping from sheet 2...")
        sheet2_data = {}
        for row in ws_sheet2.iter_rows(min_row=2, values_only=True):
            if row[0]:  # If File Reference exists
                file_ref = str(row[0]).strip()
                # Get values from columns B onwards
                values = row[1:len(sheet2_headers)+1]
                sheet2_data[file_ref] = values
        
        print(f"Mapped {len(sheet2_data)} file references from sheet 2")
        
        # Match and fill data based on File Reference (column G in template)
        print("Matching references and filling data...")
        matches_found = 0
        for row_idx in range(2, 2 + len(sheet1_data)):
            # Get the File Reference from column G (column 7)
            file_ref = ws_template.cell(row=row_idx, column=7).value
            
            if file_ref:
                file_ref = str(file_ref).strip()
                
                # If we find a match in sheet2_data
                if file_ref in sheet2_data:
                    matches_found += 1
                    values = sheet2_data[file_ref]
                    
                    # Write values starting from column I (column 9)
                    # Paste only values to preserve template formatting
                    for col_offset, value in enumerate(values):
                        ws_template.cell(row=row_idx, column=template_col_start + col_offset).value = value
        
        print(f"Matched and filled {matches_found} references")
        
        # Create output directory if it doesn't exist
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # Save the combined workbook
        print(f"Saving combined workbook to: {output_path}")
        wb_template.save(output_path)
        
        # Close workbooks
        wb_template.close()
        wb_sheet1.close()
        wb_sheet2.close()
        
        print("Successfully combined Excel sheets!")
        return output_path
        
    except FileNotFoundError as e:
        print(f"Error: File not found - {e}")
        return None
    except Exception as e:
        print(f"Error occurred: {e}")
        import traceback
        traceback.print_exc()
        return None


if __name__ == "__main__":
    # This allows testing the module directly
    print("This module should be imported and used by another script.")
    print("Please use main.py to run the Excel combiner.")