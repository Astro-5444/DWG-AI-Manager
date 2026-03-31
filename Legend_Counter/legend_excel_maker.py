import os
import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

IMAGE_EXTENSIONS = ["png", "jpg", "jpeg", "webp"]


def find_symbol_image(symbol_folder, symbol_name):
    """Find symbol image inside folder"""
    for ext in IMAGE_EXTENSIONS:
        path = os.path.join(symbol_folder, f"{symbol_name}.{ext}")
        if os.path.exists(path):
            return path
    return None


def create_excel_from_data(file_json_data, symbols_folder, output_excel):
    """
    Generate Excel file directly from data without intermediate JSON

    Parameters
    ----------
    file_json_data : list - List of dictionaries containing File Reference, Symbol Name, Count
    symbols_folder : str - Path to folder containing symbol images
    output_excel : str - Output Excel file path
    """

    # Extract unique references and symbols from the data
    file_refs = sorted(set(item["File Reference"] for item in file_json_data))
    symbols = sorted(set(item["Symbol Name"] for item in file_json_data))

    # Pivot counts
    table = defaultdict(lambda: defaultdict(int))

    for item in file_json_data:
        table[item["File Reference"]][item["Symbol Name"]] = item["Count"]

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Symbol Count"

    # Create header row with "File Reference" first, then symbols (as placeholders for images)
    headers = ["File Reference"] + symbols
    ws.append(headers)

    # Insert images in header row (over the symbol names, then clear the text)
    for col_idx, symbol in enumerate(symbols, start=2):  # Start from column B (index 2)
        img_path = find_symbol_image(symbols_folder, symbol)
        
        if img_path:
            try:
                img = ExcelImage(img_path)
                img.width = 80
                img.height = 80
                
                # Calculate cell position for image (header row, corresponding column)
                col_letter = get_column_letter(col_idx)
                ws.add_image(img, f'{col_letter}1')
                
                # Clear the cell text so only the image remains
                ws.cell(row=1, column=col_idx, value="")
            except Exception as e:
                print(f"Could not add image for {symbol}: {e}")
                # If image fails, keep the symbol name as fallback
                ws.cell(row=1, column=col_idx, value=symbol)
        else:
            # No image found, keep the symbol name as fallback
            ws.cell(row=1, column=col_idx, value=symbol)

    # Insert data rows
    for ref in file_refs:
        row = [ref]
        for symbol in symbols:
            row.append(table[ref].get(symbol, 0))
        ws.append(row)

    # Adjust column widths to accommodate images
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12  # Adjust as needed

    # Adjust row height for header to fit images
    ws.row_dimensions[1].height = 60  # Adjust as needed

    wb.save(output_excel)
    print(f"Excel saved → {output_excel}")


# Example usage would be passing the data directly instead of file path

    
"""
from legend_excel_maker import create_excel_from_json

create_excel_from_json(
    json_path="path/to/json",
    symbols_folder="path/to/symbols_folder",
    output_excel="result.xlsx"
)
"""