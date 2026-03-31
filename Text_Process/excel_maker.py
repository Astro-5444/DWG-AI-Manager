from openpyxl import Workbook
from openpyxl.styles import Font

ORDER = [
    "Project Name",
    "Drawing Title",
    "Floor",
    "Block",
    "ISSUE DATE",
    "REV.",
    "FILE REFERENCE",
]

def clean(value):
    """Clean and normalize field values."""
    if not value:
        return ""
    value = value.strip()
    if value.lower() == "not mentioned":
        return ""
    return value

def build_excel_from_txt(input_file, output_file):
    """
    Build an Excel file from a structured text file.
    
    Args:
        input_file (str): Path to the input .txt file
        output_file (str): Path to the output .xlsx file
    
    Returns:
        str: Path to the created Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Drawings"
    
    # Add headers
    ws.append(ORDER)
    
    current = {}
    skip_block = False
    row_num = 2  # Start at row 2 (after header)
    dwg_path = ""  # Track DWG PATH separately
    
    with open(input_file, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            line = line.strip()
            
            if line.startswith("FILE:"):
                if current and not skip_block:
                    # Add row data (without DWG PATH column)
                    ws.append([current.get(k, "") for k in ORDER])
                    
                    # Add hyperlink to FILE REFERENCE cell (column E)
                    file_ref = current.get("FILE REFERENCE", "")
                    
                    if file_ref and dwg_path:
                        cell = ws.cell(row=row_num, column=7)  # Column G (FILE REFERENCE)
                        cell.hyperlink = dwg_path
                        cell.value = file_ref
                        cell.font = Font(color="0563C1", underline="single")  # Blue and underlined
                    
                    row_num += 1
                
                current = {}
                skip_block = False
                dwg_path = ""
                continue
            
            if line.startswith("ERROR:"):
                skip_block = True
                continue
            
            # Check for DWG PATH but don't add to current dict
            if line.startswith("File PATH:"):
                dwg_path = clean(line.split(":", 1)[1])
                continue
            
            # Process other fields
            for key in ORDER:
                if line.startswith(key + ":"):
                    value = line.split(":", 1)[1]
                    current[key] = clean(value)
        
        # Last block
        if current and not skip_block:
            ws.append([current.get(k, "") for k in ORDER])
            
            file_ref = current.get("FILE REFERENCE", "")
            
            if file_ref and dwg_path:
                cell = ws.cell(row=row_num, column=7)
                cell.hyperlink = dwg_path
                cell.value = file_ref
                cell.font = Font(color="0563C1", underline="single")
    
    wb.save(output_file)
    print(f"✓ Excel created: {output_file}")
    return output_file

r'''
# Example usage when run as script
INPUT_FILE = "output/results.txt"
OUTPUT_FILE = "Drawings_Index.xlsx"
    
build_excel_from_txt(INPUT_FILE, OUTPUT_FILE)
'''