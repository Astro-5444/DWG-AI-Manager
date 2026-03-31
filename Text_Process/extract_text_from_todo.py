import json
from pathlib import Path
import fitz
import docx
from openpyxl import load_workbook

def extract_pdf_text(file_path):
    text = ""
    doc = fitz.open(file_path)
    for page in doc:
        text += page.get_text()
    return text.strip()

def extract_docx_text(file_path):
    doc = docx.Document(file_path)
    return "\n".join(p.text for p in doc.paragraphs).strip()

def extract_excel_text(file_path):
    wb = load_workbook(file_path, data_only=True)
    text_lines = []
    for sheet in wb:
        for row in sheet.iter_rows(values_only=True):
            row_text = [str(cell) for cell in row if cell is not None]
            if row_text:
                text_lines.append(" | ".join(row_text))
    return "\n".join(text_lines).strip()


def extract_from_todo(todo_json_path, base_output_folder=".", progress_callback=None):
    todo_json_path = Path(todo_json_path)
    with open(todo_json_path, "r", encoding="utf-8") as f:
        todo_list = json.load(f)

    # Use the TODO list name (folders) to create subfolder
    todo_base_name = todo_json_path.stem.replace(" to do list", "")
    output_folder = Path(base_output_folder) / todo_base_name
    output_folder.mkdir(parents=True, exist_ok=True)

    for i, item in enumerate(todo_list):
        if progress_callback:
            progress_callback(i + 1, len(todo_list), Path(item["path"]).name)

        file_path = Path(item["path"])
        ext = file_path.suffix.lower()

        try:
            if ext == ".pdf":
                content = extract_pdf_text(file_path)
            elif ext in {".docx", ".doc"}:
                content = extract_docx_text(file_path)
            elif ext in {".xlsx", ".xls"}:
                content = extract_excel_text(file_path)
            else:
                content = "[Unsupported file type]"

            # Name the .txt like the original file
            txt_name = f"{file_path.stem}.txt"
            txt_path = output_folder / txt_name

            with open(txt_path, "w", encoding="utf-8") as txt:
                txt.write(f"FILE REFERENCE:\n{file_path.stem}\n\n")
                txt.write(f"ORIGINAL PATH:\n{file_path}\n\n")
                txt.write("-" * 40 + "\n\n")
                txt.write("CONTENT:\n")
                txt.write(content)

            item["status"] = "Text Extracted"

        except Exception as e:
            item["status"] = "Text Extraction Failed"
            print(f"Failed to process {file_path.name}: {e}")

    # Update status back to JSON
    with open(todo_json_path, "w", encoding="utf-8") as f:
        json.dump(todo_list, f, indent=2)
